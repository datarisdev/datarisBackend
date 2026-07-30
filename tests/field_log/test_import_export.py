"""Importación desde Excel y exportación al mismo formato."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.modules.field_log.export_xlsx import build_workbook, default_file_name
from app.modules.field_log.import_xlsx import parse_workbook


def _spreadsheet_like_the_original() -> bytes:
    """Reproduce el layout de la bitácora del CDT en lo que lee el importador."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro"

    ws.cell(row=2, column=2, value="BITÁCORA CICLO:    CDT FIRA VILLADIEGO")
    ws.cell(row=4, column=2, value="RENDIMIENTO POR HA")
    ws.cell(row=4, column=4, value="PRECIO DE VENTA")
    ws.cell(row=5, column=2, value=8.4)
    ws.cell(row=5, column=4, value=5050)

    ws.cell(row=18, column=2, value="DESCRIPCION DE CONCEPTOS")
    ws.cell(row=18, column=4, value="UNIDADES")
    ws.cell(row=18, column=5, value="CANTIDAD")
    ws.cell(row=18, column=6, value="C.U. ")
    ws.cell(row=18, column=7, value="COSTO / HA ($)")

    ws.cell(row=19, column=2, value="1. ACONDICIONAMIENTO")
    ws.cell(row=20, column=2, value="Subsoleo")
    ws.cell(row=20, column=4, value="ha")
    ws.cell(row=20, column=5, value=1)
    ws.cell(row=20, column=6, value=1800)
    ws.cell(row=20, column=7, value=1800)
    ws.cell(row=20, column=9, value=date(2026, 3, 12))
    ws.cell(row=20, column=11, value="Suelo con buena humedad")

    ws.cell(row=30, column=2, value="2. SIEMBRA")
    ws.cell(row=31, column=2, value="Semilla híbrida")
    ws.cell(row=31, column=4, value="kg")
    ws.cell(row=31, column=5, value=22)
    ws.cell(row=31, column=6, value=95)
    ws.cell(row=31, column=7, value=2090)

    ws.cell(row=41, column=2, value="3. RIEGOS")
    ws.cell(row=42, column=2, value="Riego de presiembra")
    ws.cell(row=42, column=5, value=1)
    ws.cell(row=42, column=6, value=650)
    ws.cell(row=42, column=7, value=650)

    # Fenología en su columna lateral.
    ws.cell(row=26, column=13, value="Etapa")
    ws.cell(row=27, column=13, value="Germinación")
    ws.cell(row=27, column=14, value=date(2026, 3, 20))
    ws.cell(row=27, column=15, value="Emergencia pareja")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestImport:
    def test_reads_blocks_by_label_not_by_row_number(self):
        parsed = parse_workbook(_spreadsheet_like_the_original())
        categories = [entry["category"] for entry in parsed["entries"]]

        assert categories == ["acondicionamiento", "siembra", "riego"]
        assert parsed["entries"][0]["description"] == "Subsoleo"
        assert parsed["entries"][0]["cost_per_ha"] == 1800
        assert parsed["entries"][0]["performed_at"] == date(2026, 3, 12)
        assert parsed["entries"][0]["observations"] == "Suelo con buena humedad"

    def test_reads_cycle_header_and_phenology(self):
        parsed = parse_workbook(_spreadsheet_like_the_original())

        assert parsed["cycle"]["actual_yield_ton_ha"] == 8.4
        assert parsed["cycle"]["target_price_per_ton"] == 5050
        assert parsed["phenology"][0]["stage_label"] == "Germinación"
        assert parsed["phenology"][0]["observed_at"] == date(2026, 3, 20)

    def test_header_rows_are_not_imported_as_labor(self):
        parsed = parse_workbook(_spreadsheet_like_the_original())
        descriptions = [entry["description"] for entry in parsed["entries"]]
        assert "DESCRIPCION DE CONCEPTOS" not in descriptions

    def test_empty_sheet_warns_instead_of_failing(self):
        wb = Workbook()
        buffer = BytesIO()
        wb.save(buffer)

        parsed = parse_workbook(buffer.getvalue())

        assert parsed["entries"] == []
        assert parsed["warnings"]


class TestExport:
    def _summary(self):
        return {
            "cycle": {
                "name": "Ciclo O-I 2026",
                "parcel_name": "Lote 4",
                "crop_type": "Maíz",
                "variety": "Híbrido X",
                "area_ha": 12.5,
                "planting_date": date(2026, 3, 15),
                "harvest_date": None,
            },
            "kpis": {
                "economics": {
                    "yield_ton_ha": 6.0,
                    "price_per_ton": 5000.0,
                    "revenue_per_ha": 30000.0,
                    "investment_per_ha": 14156.0,
                    "profit_per_ha": 15844.0,
                    "benefit_cost_ratio": 2.12,
                    "break_even_ton_ha": 2.83,
                    "unit_cost_per_ton": 2359.33,
                },
                "sustainability": {
                    "water_footprint_m3_per_ton": 750.0,
                    "energy_kwh_per_ton": 300.0,
                    "diesel_l_per_ha": 45.0,
                    "insecticide_ia_g_per_ton": 60.0,
                    "herbicide_ia_g_per_ton": 120.0,
                    "fungicide_ia_g_per_ton": None,
                    "nitrogen_kg_per_ton": 40.0,
                },
                "costs": {
                    "total_cost_per_ha": 14156.0,
                    "categories": [
                        {
                            "category": "acondicionamiento",
                            "label": "1. Acondicionamiento",
                            "cost_per_ha": 1800.0,
                            "percentage": 12.7,
                            "entry_count": 1,
                        }
                    ],
                },
            },
            "phenology": [],
            "verification": {"verified": 8, "outside": 1, "unknown": 0, "total": 9},
        }

    def _sensitivity(self):
        return {
            "price_axis": [4450.0, 5650.0],
            "rows": [
                {
                    "yield_ton_ha": 4.5,
                    "cells": [
                        {"price_per_ton": 4450.0, "profit_per_ha": 5869.0},
                        {"price_per_ton": 5650.0, "profit_per_ha": 11269.0},
                    ],
                }
            ],
        }

    def test_produces_a_readable_workbook_with_the_expected_sections(self):
        content = build_workbook(
            self._summary(),
            [
                {
                    "category": "acondicionamiento",
                    "description": "Subsoleo",
                    "unit": "ha",
                    "quantity": 1,
                    "unit_cost": 1800,
                    "cost_per_ha": 1800,
                    "performed_at": date(2026, 3, 12),
                    "observations": "Suelo con buena humedad",
                    "data": {"diesel_l": 18},
                }
            ],
            self._sensitivity(),
        )

        ws = load_workbook(BytesIO(content)).active
        text = "\n".join(
            str(cell.value)
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        )

        assert "BITÁCORA CICLO: Ciclo O-I 2026" in text
        assert "RESULTADOS TÉCNICOS" in text
        assert "1. Acondicionamiento" in text
        assert "Subsoleo" in text
        assert "RESUMEN DE COSTOS" in text
        assert "ANÁLISIS DE SENSIBILIDAD" in text
        # La trazabilidad de la captura es información que la hoja manual nunca
        # pudo dar y es media razón para usar el módulo.
        assert "TRAZABILIDAD DE CAPTURA" in text

    def test_undefined_indicators_are_marked_not_zeroed(self):
        content = build_workbook(self._summary(), [], self._sensitivity())
        ws = load_workbook(BytesIO(content)).active
        values = [cell.value for row in ws.iter_rows() for cell in row]
        assert "—" in values

    def test_round_trip_export_then_import(self):
        """Lo exportado se vuelve a leer sin inventar labores.

        El archivo exportado termina con el resumen de costos y la matriz de
        sensibilidad, que son puros números: si el lector no cerrara la zona de
        labores al llegar ahí, cada fila de la matriz entraría como una labor
        fantasma y el costo del ciclo se duplicaría al reimportar.
        """
        content = build_workbook(
            self._summary(),
            [
                {
                    "category": "siembra",
                    "description": "Semilla híbrida",
                    "unit": "kg",
                    "quantity": 22,
                    "unit_cost": 95,
                    "cost_per_ha": 2090,
                    "performed_at": date(2026, 3, 15),
                    "observations": None,
                    "data": {},
                },
                {
                    "category": "riego",
                    "description": "Riego de presiembra",
                    "unit": "riego",
                    "quantity": 1,
                    "unit_cost": 650,
                    "cost_per_ha": 650,
                    "performed_at": date(2026, 3, 20),
                    "observations": None,
                    "data": {"m3": 3000},
                },
            ],
            self._sensitivity(),
        )

        parsed = parse_workbook(content)
        descriptions = [entry["description"] for entry in parsed["entries"]]

        assert descriptions == ["Semilla híbrida", "Riego de presiembra"]

    def test_file_name_is_safe_for_downloads(self):
        assert default_file_name("Ciclo O/I 2026") == "Bitacora - Ciclo O_I 2026.xlsx"


class TestBlockFootersInTheExport:
    """El archivo exportado cierra cada bloque como el formato en papel."""

    def _summary_with_blocks(self):
        base = TestExport()._summary()
        base["cycle"]["attributes"] = {"tipo_labranza": "Conservación", "cobertura_pct": 30}
        base["kpis"]["blocks"] = {
            "acondicionamiento": {"cost_per_ha": 1800.0, "diesel_l_per_ha": 39.0},
            "riego": {"cost_per_ha": 1750.0, "kwh_per_ha": 600.0, "m3_per_ha": 2300.0},
        }
        base["template"] = {
            "categories": [
                {
                    "key": "acondicionamiento",
                    "totals": [{"key": "diesel_l_per_ha", "label": "Litros de diesel/ha"}],
                    "attributes": [
                        {"name": "tipo_labranza", "label": "Tipo de labranza", "type": "select"},
                        {
                            "name": "cobertura_pct",
                            "label": "Porcentaje de cobertura",
                            "type": "number",
                            "unit": "%",
                        },
                    ],
                },
                {
                    "key": "riego",
                    "totals": [
                        {"key": "kwh_per_ha", "label": "Kwh/ha"},
                        {"key": "m3_per_ha", "label": "M cub/ha"},
                    ],
                },
            ]
        }
        return base

    def _text(self, content: bytes) -> str:
        ws = load_workbook(BytesIO(content)).active
        return "\n".join(
            str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None
        )

    def test_each_block_closes_with_its_own_totals(self):
        text = self._text(
            build_workbook(self._summary_with_blocks(), [], TestExport()._sensitivity())
        )

        assert "Litros de diesel/ha" in text
        assert "39" in text
        assert "Kwh/ha" in text
        assert "M cub/ha" in text

    def test_cycle_data_lands_under_its_own_block(self):
        """El tipo de labranza vive bajo acondicionamiento, como en la hoja."""
        text = self._text(
            build_workbook(self._summary_with_blocks(), [], TestExport()._sensitivity())
        )

        assert "Tipo de labranza" in text
        assert "Conservación" in text
        assert "Porcentaje de cobertura (%)" in text

    def test_a_template_without_declarations_exports_as_before(self):
        """Una plantilla vieja, sin pies declarados, no rompe la exportación."""
        summary = TestExport()._summary()
        text = self._text(build_workbook(summary, [], TestExport()._sensitivity()))

        assert "RESUMEN DE COSTOS" in text
