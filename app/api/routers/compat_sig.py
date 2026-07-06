from __future__ import annotations

import csv
import hashlib
import io
import json
import uuid
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from fastapi import APIRouter, Body, File, Form, Header, HTTPException, Query, UploadFile
from fastapi.concurrency import run_in_threadpool
from openpyxl import load_workbook
from shapely.geometry import Point, shape
from shapely.geometry.base import BaseGeometry

from app.api.routers.compat import LOCK, bearer_user, now, read_db, table, write_db
from app.core.config import settings
from app.services.telemetry.sig_planning import process_planning_upload
from app.services.digiforms_data_api import (
    DigiformsDataAPI,
    DigiformsDataAPIError,
    discover_field_names,
    dynamic_fields_from_payload,
    extract_image_rows,
    extract_submission_rows,
    image_path_from_payload,
    response_id_from_payload,
    state_from_payload,
)
from app.services.digiforms_company_config import (
    connection_for_company,
    configured_form_types as tenant_configured_form_types,
    mapping_for_company,
    runtime_credentials,
    safe_connection,
    safe_mappings,
    mappings_for_company,
)
from app.api.routers.compat_extensions import company_for_user, extension_enabled_for
from app.services.commercial_demo_seed import is_commercial_demo_user

router = APIRouter(prefix="/compat/sig-agricola", tags=["SIG Agrícola compatibility"])

MAX_UPLOAD_BYTES = 15 * 1024 * 1024
MAX_ROWS_PER_IMPORT = 25_000
HARVEST_FORM_TYPE = "harvest"
PEST_WEED_FORM_TYPE = "pest_weed"
DIGIFORMS_EXCEL_SOURCE = "digiforms_excel_export"
DIGIFORMS_RESULTS_API_SOURCE = "digiforms_results_api"


def _require_user(authorization: Optional[str]) -> Dict[str, Any]:
    user = bearer_user(authorization)
    if not user:
        raise HTTPException(status_code=401, detail="No autenticado")
    return user


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="minutes")
    return str(value).strip()


def _json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return _safe_text(value)
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    return value


def _normalize(value: Any) -> str:
    text = _safe_text(value).lower()
    replacements = str.maketrans("áéíóúüñ", "aeiouun")
    return "".join(ch for ch in text.translate(replacements) if ch.isalnum())


def _first_value(row: Dict[str, Any], aliases: Sequence[str], default: Any = "") -> Any:
    normalized = {_normalize(key): value for key, value in row.items() if not str(key).startswith("_")}
    for alias in aliases:
        key = _normalize(alias)
        if key in normalized and normalized[key] not in (None, ""):
            return normalized[key]
    return default


def _to_float(value: Any) -> Optional[float]:
    try:
        number = float(str(value).strip().replace(" ", ""))
    except (TypeError, ValueError):
        return None
    return number


def _parse_coordinates(row: Dict[str, Any]) -> Optional[Tuple[float, float]]:
    latitude = _first_value(row, ["latitud", "lat", "latitude"], None)
    longitude = _first_value(row, ["longitud", "lng", "lon", "longitude", "long"], None)
    if latitude not in (None, "") and longitude not in (None, ""):
        lat = _to_float(latitude)
        lng = _to_float(longitude)
        if lat is not None and lng is not None and not (lat == 0 and lng == 0):
            return lat, lng

    raw = _safe_text(
        _first_value(
            row,
            [
                "GeoLocalizacion",
                "Geo Localizacion",
                "Geolocalización",
                "UBICACION",
                "Ubicación",
                "Location",
                "coordenadas",
                "coords",
            ],
        )
    )
    if not raw or raw.replace(" ", "") in {"-1,-1", "0,0"}:
        return None
    parts = [part.strip() for part in raw.replace(";", ",").split(",")]
    if len(parts) < 2:
        return None
    first = _to_float(parts[0])
    second = _to_float(parts[1])
    if first is None or second is None:
        return None
    if abs(first) > 90:
        lng, lat = first, second
    else:
        lat, lng = first, second
    if lat == 0 and lng == 0:
        return None
    if not (-90 <= lat <= 90 and -180 <= lng <= 180):
        return None
    return lat, lng


def _header_score(values: Sequence[Any], form_type: str) -> int:
    normalized = {_normalize(value) for value in values if value not in (None, "")}
    if form_type == PEST_WEED_FORM_TYPE:
        targets = {"idrespuesta", "geolocalizacion", "maleza", "plaga", "evidenciafotografica"}
    else:
        targets = {"responseid", "ubicacion", "status", "metododecosecha", "parcela"}
    return sum(1 for target in targets if target in normalized)


def _read_xlsx_rows(content: bytes, form_type: str) -> List[Dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel .xlsx: {exc}") from exc
    worksheet = workbook[workbook.sheetnames[0]]
    raw_rows = list(worksheet.iter_rows(values_only=False))
    if not raw_rows:
        return []

    search_limit = min(len(raw_rows), 35)
    header_index = max(range(search_limit), key=lambda index: _header_score([cell.value for cell in raw_rows[index]], form_type))
    if _header_score([cell.value for cell in raw_rows[header_index]], form_type) < 2:
        raise HTTPException(status_code=400, detail="No se identificaron encabezados compatibles con la exportación de DigiForms.")

    headers = [_safe_text(cell.value) or f"col_{position + 1}" for position, cell in enumerate(raw_rows[header_index])]
    rows: List[Dict[str, Any]] = []
    for cells in raw_rows[header_index + 1 :]:
        values = [cell.value for cell in cells]
        if not any(value not in (None, "") for value in values):
            continue
        row: Dict[str, Any] = {}
        hyperlinks: Dict[str, str] = {}
        for index, header in enumerate(headers):
            cell = cells[index] if index < len(cells) else None
            value = cell.value if cell is not None else None
            row[header] = value
            if cell is not None and cell.hyperlink and cell.hyperlink.target:
                hyperlinks[_normalize(header)] = str(cell.hyperlink.target)
                hyperlinks[f"col_{index + 1}"] = str(cell.hyperlink.target)
                hyperlinks[f"value_{_normalize(value)}"] = str(cell.hyperlink.target)
        row["_hyperlinks"] = hyperlinks
        rows.append(row)
        if len(rows) >= MAX_ROWS_PER_IMPORT:
            break
    return rows


def _read_csv_rows(content: bytes) -> List[Dict[str, Any]]:
    decoded = content.decode("utf-8-sig", errors="replace")
    sample = decoded[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(decoded), dialect=dialect)
    rows = [dict(row) for row in reader if row]
    return rows[:MAX_ROWS_PER_IMPORT]


def _read_rows(filename: str, content: bytes, form_type: str) -> List[Dict[str, Any]]:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(content)
    if suffix == ".xlsx":
        return _read_xlsx_rows(content, form_type)
    if suffix == ".xls":
        raise HTTPException(status_code=400, detail="El formato .xls antiguo no es compatible. Exporta el reporte de DigiForms como .xlsx o .csv.")
    raise HTTPException(status_code=400, detail="Formato no compatible. Usa una exportación .xlsx o .csv de DigiForms.")


def _parse_parcels(raw: str) -> List[Dict[str, Any]]:
    try:
        payload = json.loads(raw or "[]")
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Las parcelas enviadas no contienen JSON válido.") from exc
    if isinstance(payload, dict):
        payload = payload.get("parcels") or payload.get("data") or []
    if not isinstance(payload, list):
        raise HTTPException(status_code=400, detail="El listado de parcelas debe ser un arreglo JSON.")
    return [parcel for parcel in payload if isinstance(parcel, dict)]


def _feature_name(properties: Dict[str, Any], fallback: str) -> str:
    for key in ["name", "Name", "NAME", "ID", "id", "lote", "Lote", "codigo", "Código"]:
        value = properties.get(key)
        if value not in (None, ""):
            return _safe_text(value)
    return fallback


def _iter_parcel_features(parcels: Sequence[Dict[str, Any]]) -> Iterable[Tuple[str, str, BaseGeometry]]:
    for parcel in parcels:
        parcel_id = _safe_text(parcel.get("id"))
        if not parcel_id:
            continue
        parcel_name = _safe_text(parcel.get("name")) or parcel_id
        geometry = parcel.get("geometry") or parcel.get("geometry_geojson")
        if not isinstance(geometry, dict):
            continue
        if geometry.get("type") == "FeatureCollection" or isinstance(geometry.get("features"), list):
            for feature in geometry.get("features") or []:
                if not isinstance(feature, dict) or not isinstance(feature.get("geometry"), dict):
                    continue
                try:
                    feature_geometry = shape(feature["geometry"])
                except Exception:
                    continue
                if feature_geometry.is_empty:
                    continue
                name = _feature_name(feature.get("properties") or {}, parcel_name)
                yield parcel_id, f"{parcel_id}::{name}", feature_geometry
            continue
        if geometry.get("type") == "Feature" and isinstance(geometry.get("geometry"), dict):
            try:
                feature_geometry = shape(geometry["geometry"])
            except Exception:
                continue
            if not feature_geometry.is_empty:
                name = _feature_name(geometry.get("properties") or {}, parcel_name)
                yield parcel_id, f"{parcel_id}::{name}", feature_geometry
            continue
        try:
            parcel_geometry = shape(geometry)
        except Exception:
            continue
        if not parcel_geometry.is_empty:
            yield parcel_id, f"{parcel_id}::{parcel_name}", parcel_geometry


def _build_feature_index(parcels: Sequence[Dict[str, Any]]) -> List[Tuple[str, str, BaseGeometry]]:
    return list(_iter_parcel_features(parcels))


def _match_point(lat: float, lng: float, features: Sequence[Tuple[str, str, BaseGeometry]]) -> Tuple[Optional[str], Optional[str]]:
    point = Point(lng, lat)
    for parcel_id, feature_key, geometry in features:
        try:
            if geometry.covers(point):
                return parcel_id, feature_key
        except Exception:
            continue
    return None, None


def _canonical_harvest_state(value: Any) -> str:
    raw = _safe_text(value).lower()
    if any(word in raw for word in ["terminad", "finaliz", "completed", "done", "cosechad"]):
        return "Terminada"
    if any(word in raw for word in ["proceso", "parcial", "progress"]):
        return "En proceso"
    return "No iniciada"


def _stable_external_id(form_type: str, row: Dict[str, Any], index: int) -> str:
    value = _first_value(row, ["Response Id", "IdRespuesta", "response_id", "id_respuesta", "id"], "")
    if value not in (None, ""):
        return _safe_text(value)
    safe_payload = json.dumps(_json_safe(row), sort_keys=True, ensure_ascii=False, default=str)
    digest = hashlib.sha256(f"{form_type}:{index}:{safe_payload}".encode("utf-8")).hexdigest()[:20]
    return f"generated-{digest}"


def _harvest_record(row: Dict[str, Any], index: int, features: Sequence[Tuple[str, str, BaseGeometry]]) -> Optional[Dict[str, Any]]:
    coords = _parse_coordinates(row)
    if not coords:
        return None
    lat, lng = coords
    parcel_id, feature_key = _match_point(lat, lng, features)
    return {
        "external_response_id": _stable_external_id(HARVEST_FORM_TYPE, row, index),
        "parcela": _safe_text(_first_value(row, ["Parcela", "Nombre", "Lote"], f"Parcela {index + 1}")),
        "zona": _safe_text(_first_value(row, ["Zona", "Bloque"], "-")) or "-",
        "metodo": _safe_text(_first_value(row, ["Metodo de Cosecha", "Metodode Cosecha", "Método", "Metodo"], "-")) or "-",
        "canacond": _safe_text(_first_value(row, ["CANACOND", "Condicion", "Condición"], "-")) or "-",
        "terminal": _safe_text(_first_value(row, ["Terminal", "Operador"], "-")) or "-",
        "estado": _canonical_harvest_state(_first_value(row, ["Status", "Estado", "State"], "")),
        "lat": lat,
        "lng": lng,
        "hora_inicio": _safe_text(_first_value(row, ["Hora", "Hora Inicio", "Start At"], "-")) or "-",
        "hora_fin": _safe_text(_first_value(row, ["Hora Finalizacion", "Hora Finalización", "End At"], "-")) or "-",
        "fecha": _safe_text(_first_value(row, ["Fecha"], "-")) or "-",
        "fecha_fin": _safe_text(_first_value(row, ["Fecha Finalizacion", "Fecha Finalización", "Fecha Fin"], "-")) or "-",
        "parcel_id": parcel_id,
        "feature_key": feature_key,
        "outside_registered_parcel": feature_key is None,
        "digiforms_state": _safe_text(_first_value(row, ["State", "EstadoAprobacion", "ApprovalState"], "0")) or "0",
        "image_urls": [_safe_text(value) for value in (row.get("_image_urls") or []) if _safe_text(value)],
        "dynamic_fields": _json_safe(dynamic_fields_from_payload(row)),
        "raw_payload": _json_safe(row.get("_raw_api_payload") or {key: value for key, value in row.items() if key not in {"_hyperlinks", "_raw_api_payload"}}),
    }


def _photo_links(row: Dict[str, Any]) -> Tuple[str, str]:
    hyperlinks = row.get("_hyperlinks") or {}
    if not isinstance(hyperlinks, dict):
        hyperlinks = {}
    api_image_urls = [
        _safe_text(item)
        for item in (row.get("_image_urls") or [])
        if _safe_text(item)
    ]
    evidence = api_image_urls[0] if api_image_urls else ""
    signature = api_image_urls[1] if len(api_image_urls) > 1 else ""
    for key, value in hyperlinks.items():
        normalized_key = _normalize(key)
        if not value:
            continue
        if "firma" in normalized_key:
            signature = signature or _safe_text(value)
        elif "evidencia" in normalized_key or "fotograf" in normalized_key:
            evidence = evidence or _safe_text(value)
    if not evidence:
        # DigiForms exports may put the evidence image immediately after the text column.
        evidence = _safe_text(hyperlinks.get("col_23") or hyperlinks.get("col_22") or "")
    if not signature:
        signature = _safe_text(hyperlinks.get("col_24") or hyperlinks.get("col_25") or "")
    return evidence, signature


def _pest_weed_record(row: Dict[str, Any], index: int, features: Sequence[Tuple[str, str, BaseGeometry]]) -> Optional[Dict[str, Any]]:
    coords = _parse_coordinates(row)
    if not coords:
        return None
    lat, lng = coords
    parcel_id, feature_key = _match_point(lat, lng, features)
    photo_url, signature_url = _photo_links(row)
    return {
        "external_response_id": _stable_external_id(PEST_WEED_FORM_TYPE, row, index),
        "terminal": _safe_text(_first_value(row, ["Terminal"], "")),
        "fecha": _safe_text(_first_value(row, ["Fecha"], "")),
        "hora": _safe_text(_first_value(row, ["Hora"], "")),
        "zona": _safe_text(_first_value(row, ["Zona"], "")),
        "maleza": _safe_text(_first_value(row, ["Maleza"], "")),
        "tipo_maleza": _safe_text(_first_value(row, ["TipodeMaleza", "Tipo de Maleza"], "")),
        "plaga": _safe_text(_first_value(row, ["Plaga"], "")),
        "tipo_plaga": _safe_text(_first_value(row, ["TiposdePlaga", "Tipos de Plaga", "Tipo de Plaga"], "")),
        "enfermedades": _safe_text(_first_value(row, ["Enfermedades"], "")),
        "sintomas": _safe_text(_first_value(row, ["Sintomas", "Síntomas"], "")),
        "observaciones": _safe_text(_first_value(row, ["Observaciones"], "")),
        "firma": signature_url,
        "firma_texto": _safe_text(_first_value(row, ["Firma"], "")),
        "evidencia_fotografica": _safe_text(_first_value(row, ["EvidenciaFotografica", "Evidencia Fotografica", "Evidencia Fotográfica"], "")),
        "photo_url": photo_url,
        "signature_url": signature_url,
        "lat": lat,
        "lng": lng,
        "parcel_id": parcel_id,
        "feature_key": feature_key,
        "outside_registered_parcel": feature_key is None,
        "digiforms_state": _safe_text(_first_value(row, ["State", "EstadoAprobacion", "ApprovalState"], "0")) or "0",
        "image_urls": [_safe_text(value) for value in (row.get("_image_urls") or []) if _safe_text(value)],
        "dynamic_fields": _json_safe(dynamic_fields_from_payload(row)),
        "raw_payload": _json_safe(row.get("_raw_api_payload") or {key: value for key, value in row.items() if key not in {"_hyperlinks", "_raw_api_payload"}}),
    }


def _record_table_name(form_type: str) -> str:
    return "sig_harvest_records" if form_type == HARVEST_FORM_TYPE else "sig_pest_weed_records"


def _safe_record(row: Dict[str, Any]) -> Dict[str, Any]:
    return dict(row)


def _same_company_or_legacy_user(row: Dict[str, Any], *, company_id: Optional[str], user_id: str) -> bool:
    row_company_id = str(row.get("company_id") or "")
    if company_id and row_company_id:
        return row_company_id == str(company_id)
    return str(row.get("user_id") or "") == user_id


def _import_records(
    *,
    db: Dict[str, Any],
    user_id: str,
    company_id: Optional[str],
    form_type: str,
    filename: str,
    content: bytes,
    parcels: Sequence[Dict[str, Any]],
) -> Dict[str, Any]:
    rows = _read_rows(filename, content, form_type)
    features = _build_feature_index(parcels)
    parser = _harvest_record if form_type == HARVEST_FORM_TYPE else _pest_weed_record
    parsed_records: List[Dict[str, Any]] = []
    invalid_rows = 0
    for index, row in enumerate(rows):
        record = parser(row, index, features)
        if record is None:
            invalid_rows += 1
            continue
        parsed_records.append(record)

    if not parsed_records:
        raise HTTPException(status_code=400, detail="No se encontraron registros con coordenadas válidas en la exportación.")

    import_run_id = str(uuid.uuid4())
    timestamp = now()
    destination = table(db, _record_table_name(form_type))
    existing_by_external = {
        str(item.get("external_response_id")): item
        for item in destination
        if _same_company_or_legacy_user(item, company_id=company_id, user_id=user_id) and item.get("external_response_id")
    }
    imported_rows = 0
    updated_rows = 0
    duplicate_rows = 0
    seen_in_file: set[str] = set()

    for parsed in parsed_records:
        external_id = str(parsed["external_response_id"])
        if external_id in seen_in_file:
            duplicate_rows += 1
            continue
        seen_in_file.add(external_id)
        common = {
            **parsed,
            "user_id": user_id,
            "company_id": company_id,
            "source": DIGIFORMS_EXCEL_SOURCE,
            "source_file_name": filename,
            "import_run_id": import_run_id,
            "updated_at": timestamp,
        }
        existing = existing_by_external.get(external_id)
        if existing:
            existing.update(common)
            updated_rows += 1
        else:
            item = {"id": str(uuid.uuid4()), "created_at": timestamp, **common}
            destination.append(item)
            existing_by_external[external_id] = item
            imported_rows += 1

    outside_rows = sum(1 for item in parsed_records if item.get("outside_registered_parcel"))
    run = {
        "id": import_run_id,
        "user_id": user_id,
        "company_id": company_id,
        "form_type": form_type,
        "source": DIGIFORMS_EXCEL_SOURCE,
        "file_name": filename,
        "status": "completed",
        "total_rows": len(rows),
        "valid_rows": len(parsed_records),
        "imported_rows": imported_rows,
        "updated_rows": updated_rows,
        "duplicate_rows": duplicate_rows,
        "invalid_rows": invalid_rows,
        "outside_registered_parcel_rows": outside_rows,
        "parcel_features_received": len(features),
        "created_at": timestamp,
        "updated_at": timestamp,
    }
    table(db, "sig_import_runs").append(run)
    return run


async def _read_upload(file: UploadFile) -> bytes:
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="El archivo está vacío.")
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="El archivo supera el límite de 15 MB.")
    return content


def _company_id_for_user(db: Dict[str, Any], user_id: str) -> Optional[str]:
    return company_for_user(db, user_id)


def _form_id_for_type(db: Dict[str, Any], company_id: Optional[str], form_type: str) -> str:
    mapping = mapping_for_company(db, company_id, form_type)
    if mapping and mapping.get("is_enabled", True) is not False:
        return _safe_text(mapping.get("form_id"))
    # Backward-compatible fallback for deployments configured before the
    # multi-company flow existed.
    if form_type == HARVEST_FORM_TYPE:
        return _safe_text(settings.DIGIFORMS_HARVEST_FORM_ID)
    if form_type == PEST_WEED_FORM_TYPE:
        return _safe_text(settings.DIGIFORMS_PEST_WEED_FORM_ID)
    return ""


def _initial_response_id_for_type(db: Dict[str, Any], company_id: Optional[str], form_type: str) -> int:
    mapping = mapping_for_company(db, company_id, form_type)
    if mapping:
        return int(mapping.get("initial_response_id") or 0)
    return int(settings.DIGIFORMS_SYNC_INITIAL_RESPONSE_ID or 0)


def _configured_form_types(db: Dict[str, Any], company_id: Optional[str]) -> List[str]:
    configured = tenant_configured_form_types(db, company_id)
    if configured:
        return configured
    return [form_type for form_type in [HARVEST_FORM_TYPE, PEST_WEED_FORM_TYPE] if _form_id_for_type(db, company_id, form_type)]


def _initial_sync_dates_for_type(db: Dict[str, Any], company_id: Optional[str], form_type: str) -> Tuple[str, str]:
    mapping = mapping_for_company(db, company_id, form_type) or {}
    current_year = date.today().year
    start_date = _safe_text(mapping.get("initial_sync_start_date")) or f"{current_year}-01-01"
    end_date = _safe_text(mapping.get("initial_sync_end_date")) or f"{current_year}-12-31"
    return start_date, end_date


def _company_user_ids(db: Dict[str, Any], company_id: Optional[str]) -> set[str]:
    if not company_id:
        return set()
    ids = {
        str(row.get("user_id") or "")
        for row in table(db, "admin_users")
        if str(row.get("company_id") or "") == str(company_id) and row.get("is_active", True) is not False and row.get("user_id")
    }
    ids.update(
        str(row.get("user_id") or row.get("id") or "")
        for row in table(db, "profiles")
        if str(row.get("company_id") or "") == str(company_id) and (row.get("user_id") or row.get("id"))
    )
    company = next((row for row in table(db, "companies") if str(row.get("id") or "") == str(company_id)), None)
    if company and company.get("name"):
        ids.update(
            str(row.get("user_id") or row.get("id") or "")
            for row in table(db, "profiles")
            if row.get("company_name") == company.get("name") and (row.get("user_id") or row.get("id"))
        )
    return {value for value in ids if value}


def _stored_parcels(db: Dict[str, Any], user_id: str, company_id: Optional[str] = None) -> List[Dict[str, Any]]:
    allowed_user_ids = _company_user_ids(db, company_id)
    allowed_user_ids.add(user_id)
    return [dict(row) for row in table(db, "parcels") if str(row.get("user_id") or "") in allowed_user_ids]


def _data_api_for_company(db: Dict[str, Any], company_id: Optional[str]) -> DigiformsDataAPI:
    connection = connection_for_company(db, company_id)
    if connection:
        try:
            credentials = runtime_credentials(connection)
        except RuntimeError as exc:
            raise HTTPException(status_code=503, detail=str(exc)) from exc
        return DigiformsDataAPI(
            client_id=credentials.get("client_id"),
            api_user=credentials.get("api_user"),
            api_password=credentials.get("api_password"),
        )
    return DigiformsDataAPI()


def _cursor_row(
    db: Dict[str, Any],
    company_id: Optional[str],
    user_id: str,
    form_type: str,
    *,
    create: bool = False,
) -> Optional[Dict[str, Any]]:
    rows = table(db, "sig_sync_cursors")
    row = next(
        (
            item
            for item in rows
            if item.get("form_type") == form_type
            and (
                (company_id and str(item.get("company_id") or "") == str(company_id))
                or (not company_id and str(item.get("user_id") or "") == user_id)
            )
        ),
        None,
    )
    if row is None and create:
        timestamp = now()
        row = {
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "user_id": user_id,
            "form_type": form_type,
            "form_id": _form_id_for_type(db, company_id, form_type),
            "last_response_id": _initial_response_id_for_type(db, company_id, form_type),
            "last_sync_at": None,
            "last_success_at": None,
            "last_error": None,
            "created_at": timestamp,
            "updated_at": timestamp,
        }
        rows.append(row)
    elif row is not None:
        row.setdefault("company_id", company_id)
        row.setdefault("user_id", user_id)
    return row


def _numeric_response_id(value: Any) -> Optional[int]:
    raw = _safe_text(value)
    if not raw:
        return None
    try:
        return int(float(raw))
    except (TypeError, ValueError):
        digits = "".join(ch for ch in raw if ch.isdigit())
        return int(digits) if digits else None


def _highest_response_id(rows: Sequence[Dict[str, Any]], default: int) -> int:
    values = [_numeric_response_id(response_id_from_payload(row)) for row in rows]
    return max([value for value in values if value is not None], default=default)


def _api_image_map(image_rows: Sequence[Dict[str, Any]]) -> Dict[str, List[str]]:
    by_response: Dict[str, List[str]] = {}
    for image in image_rows:
        response_id = response_id_from_payload(image)
        image_path = image_path_from_payload(image)
        if not response_id or not image_path:
            continue
        urls = by_response.setdefault(response_id, [])
        if image_path not in urls:
            urls.append(image_path)
    return by_response


def _state_stats(records: Sequence[Dict[str, Any]]) -> Dict[str, int]:
    stats = {"approved_rows": 0, "pending_rows": 0, "rejected_rows": 0, "deleted_rows": 0, "other_state_rows": 0}
    for record in records:
        state = _safe_text(record.get("digiforms_state")) or "0"
        if state == "0":
            stats["approved_rows"] += 1
        elif state == "8":
            stats["pending_rows"] += 1
        elif state == "9":
            stats["rejected_rows"] += 1
        elif state == "5":
            stats["deleted_rows"] += 1
        else:
            stats["other_state_rows"] += 1
    return stats


def _persist_raw_api_submissions(
    *,
    db: Dict[str, Any],
    user_id: str,
    company_id: Optional[str],
    form_type: str,
    form_id: str,
    response_rows: Sequence[Dict[str, Any]],
    image_map: Dict[str, List[str]],
    timestamp: str,
) -> int:
    """Persist every variable DigiForms submission before specialized parsing.

    A configured FormId can add, remove or rename questions at any moment. Raw
    storage guarantees that no provider data is lost when a new field is not yet
    represented by a SIG card.
    """
    destination = table(db, "sig_digiforms_raw_submissions")
    existing = {
        (str(item.get("form_id") or ""), str(item.get("external_response_id") or "")): item
        for item in destination
        if _same_company_or_legacy_user(item, company_id=company_id, user_id=user_id)
    }
    persisted = 0
    for index, row in enumerate(response_rows[:MAX_ROWS_PER_IMPORT]):
        external_id = response_id_from_payload(row) or _stable_external_id(form_type, row, index)
        values = {
            "user_id": user_id,
            "company_id": company_id,
            "form_type": form_type,
            "form_id": form_id,
            "external_response_id": external_id,
            "digiforms_state": state_from_payload(row),
            "dynamic_fields": _json_safe(dynamic_fields_from_payload(row)),
            "raw_payload": _json_safe(row.get("_raw_api_payload") or row),
            "image_urls": list(image_map.get(external_id, [])),
            "updated_at": timestamp,
        }
        current = existing.get((form_id, external_id))
        if current:
            current.update(values)
        else:
            current = {"id": str(uuid.uuid4()), "created_at": timestamp, **values}
            destination.append(current)
            existing[(form_id, external_id)] = current
        persisted += 1
    return persisted


def _persist_api_records(
    *,
    db: Dict[str, Any],
    user_id: str,
    company_id: Optional[str],
    form_type: str,
    form_id: str,
    response_rows: Sequence[Dict[str, Any]],
    image_rows: Sequence[Dict[str, Any]],
    parcels: Sequence[Dict[str, Any]],
    cursor_before: int,
    sync_mode: str,
    initial_sync_start_date: Optional[str] = None,
    initial_sync_end_date: Optional[str] = None,
    warning: Optional[str] = None,
) -> Dict[str, Any]:
    features = _build_feature_index(parcels)
    image_map = _api_image_map(image_rows)
    parser = _harvest_record if form_type == HARVEST_FORM_TYPE else _pest_weed_record
    parsed_records: List[Dict[str, Any]] = []
    invalid_rows = 0

    for index, api_row in enumerate(response_rows[:MAX_ROWS_PER_IMPORT]):
        row = dict(api_row)
        response_id = response_id_from_payload(row)
        row["_image_urls"] = image_map.get(response_id, [])
        row.setdefault("State", state_from_payload(row))
        record = parser(row, index, features)
        if record is None:
            invalid_rows += 1
            continue
        record["digiforms_form_id"] = form_id
        parsed_records.append(record)

    run_id = str(uuid.uuid4())
    timestamp = now()
    raw_rows_persisted = _persist_raw_api_submissions(
        db=db,
        user_id=user_id,
        company_id=company_id,
        form_type=form_type,
        form_id=form_id,
        response_rows=response_rows,
        image_map=image_map,
        timestamp=timestamp,
    )
    destination = table(db, _record_table_name(form_type))
    existing_by_external = {
        str(item.get("external_response_id")): item
        for item in destination
        if _same_company_or_legacy_user(item, company_id=company_id, user_id=user_id) and item.get("external_response_id")
    }
    imported_rows = 0
    updated_rows = 0
    duplicate_rows = 0
    seen_in_response: set[str] = set()

    for parsed in parsed_records:
        external_id = str(parsed["external_response_id"])
        if external_id in seen_in_response:
            duplicate_rows += 1
            continue
        seen_in_response.add(external_id)
        common = {
            **parsed,
            "user_id": user_id,
            "company_id": company_id,
            "source": DIGIFORMS_RESULTS_API_SOURCE,
            "source_file_name": "",
            "import_run_id": run_id,
            "updated_at": timestamp,
        }
        existing = existing_by_external.get(external_id)
        if existing:
            existing.update(common)
            updated_rows += 1
        else:
            item = {"id": str(uuid.uuid4()), "created_at": timestamp, **common}
            destination.append(item)
            existing_by_external[external_id] = item
            imported_rows += 1

    cursor_after = _highest_response_id(response_rows, cursor_before)
    cursor = _cursor_row(db, company_id, user_id, form_type, create=True)
    if cursor is not None:
        cursor.update(
            {
                "form_id": form_id,
                "last_response_id": cursor_after,
                "last_sync_at": timestamp,
                "last_success_at": timestamp,
                "last_error": None,
                "last_request_mode": sync_mode,
                "updated_at": timestamp,
            }
        )

    outside_rows = sum(1 for item in parsed_records if item.get("outside_registered_parcel"))
    run = {
        "id": run_id,
        "user_id": user_id,
        "company_id": company_id,
        "form_type": form_type,
        "form_id": form_id,
        "source": DIGIFORMS_RESULTS_API_SOURCE,
        "file_name": "API por fechas" if sync_mode == "initial_date_range" else "API incremental",
        "status": "completed",
        "total_rows": len(response_rows),
        "valid_rows": len(parsed_records),
        "imported_rows": imported_rows,
        "updated_rows": updated_rows,
        "duplicate_rows": duplicate_rows,
        "invalid_rows": invalid_rows,
        "outside_registered_parcel_rows": outside_rows,
        "parcel_features_received": len(features),
        "images_received": len(image_rows),
        "raw_rows_persisted": raw_rows_persisted,
        "discovered_fields": discover_field_names(response_rows),
        "sync_mode": sync_mode,
        "initial_sync_start_date": initial_sync_start_date,
        "initial_sync_end_date": initial_sync_end_date,
        "cursor_before": cursor_before,
        "cursor_after": cursor_after,
        "warning": warning,
        **_state_stats(parsed_records),
        "created_at": timestamp,
        "updated_at": timestamp,
    }
    table(db, "sig_import_runs").append(run)
    return run


async def _sync_form_from_api(
    *,
    company_id: Optional[str],
    user_id: str,
    form_type: str,
    parcels: Sequence[Dict[str, Any]],
    from_response_id: Optional[int] = None,
) -> Dict[str, Any]:
    with LOCK:
        db = read_db()
        form_id = _form_id_for_type(db, company_id, form_type)
        if not form_id:
            raise HTTPException(status_code=400, detail=f"No se configuró FormId para {form_type}.")
        api = _data_api_for_company(db, company_id)
        if not api.is_configured:
            raise HTTPException(status_code=400, detail="DigiForms Data API no está configurada para esta empresa.")
        cursor = _cursor_row(db, company_id, user_id, form_type, create=True)
        cursor_before = int(from_response_id if from_response_id is not None else (cursor or {}).get("last_response_id") or 0)
        write_db(db)

    with LOCK:
        db = read_db()
        initial_start_date, initial_end_date = _initial_sync_dates_for_type(db, company_id, form_type)
    sync_mode = "incremental_response_id" if cursor_before > 0 else "initial_date_range"
    try:
        if sync_mode == "incremental_response_id":
            results_payload = await api.get_all_results_since(form_id, cursor_before)
        else:
            results_payload = await api.get_all_results_by_dates(form_id, initial_start_date, initial_end_date)
        response_rows = extract_submission_rows(results_payload)
    except DigiformsDataAPIError as exc:
        with LOCK:
            db = read_db()
            cursor = _cursor_row(db, company_id, user_id, form_type, create=True)
            if cursor is not None:
                cursor.update({"last_sync_at": now(), "last_error": str(exc), "updated_at": now()})
            write_db(db)
        raise HTTPException(status_code=502, detail=str(exc)) from exc

    image_rows: List[Dict[str, Any]] = []
    image_warning: Optional[str] = None
    try:
        if sync_mode == "incremental_response_id":
            image_payload = await api.get_images_since(form_id, cursor_before)
        else:
            image_payload = await api.get_images_by_dates(form_id, initial_start_date, initial_end_date)
        image_rows = extract_image_rows(image_payload)
    except DigiformsDataAPIError as exc:
        image_warning = f"Los registros se sincronizaron, pero no se pudieron actualizar imágenes: {exc}"

    with LOCK:
        db = read_db()
        run = _persist_api_records(
            db=db,
            user_id=user_id,
            company_id=company_id,
            form_type=form_type,
            form_id=form_id,
            response_rows=response_rows,
            image_rows=image_rows,
            parcels=parcels,
            cursor_before=cursor_before,
            sync_mode=sync_mode,
            initial_sync_start_date=initial_start_date if sync_mode == "initial_date_range" else None,
            initial_sync_end_date=initial_end_date if sync_mode == "initial_date_range" else None,
            warning=image_warning,
        )
        write_db(db)
    return run


def _sync_payload_parcels(payload: Dict[str, Any], user_id: str, company_id: Optional[str]) -> List[Dict[str, Any]]:
    parcels = payload.get("parcels")
    if isinstance(parcels, list):
        return [parcel for parcel in parcels if isinstance(parcel, dict)]
    with LOCK:
        return _stored_parcels(read_db(), user_id, company_id)


def _sync_types(db: Dict[str, Any], company_id: Optional[str], payload: Dict[str, Any]) -> List[str]:
    requested = _safe_text(payload.get("form_type") or "all")
    if requested in {HARVEST_FORM_TYPE, PEST_WEED_FORM_TYPE}:
        return [requested]
    return _configured_form_types(db, company_id)


def _cursor_public(row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "form_type": row.get("form_type"),
        "form_id": row.get("form_id"),
        "last_response_id": row.get("last_response_id", 0),
        "last_sync_at": row.get("last_sync_at"),
        "last_success_at": row.get("last_success_at"),
        "last_error": row.get("last_error"),
        "last_request_mode": row.get("last_request_mode"),
    }


@router.get("/integration-status")
def integration_status(authorization: Optional[str] = Header(default=None)):
    user = _require_user(authorization)
    user_id = str(user.get("id") or "")
    with LOCK:
        db = read_db()
        company_id = _company_id_for_user(db, user_id)
        if is_commercial_demo_user(user):
            imports = [row for row in table(db, "sig_import_runs") if _same_company_or_legacy_user(row, company_id=company_id, user_id=user_id)]
            latest = max(imports, key=lambda row: str(row.get("created_at") or ""), default=None)
            cursors = [_cursor_public(row) for row in table(db, "sig_sync_cursors") if _same_company_or_legacy_user(row, company_id=company_id, user_id=user_id)]
            mappings = safe_mappings(mappings_for_company(db, company_id))
            return {
                "data": {
                    "company_id": company_id,
                    "capture_platform": "DigiformsApp",
                    "active_results_channel": "commercial_demo_dataset",
                    "active_results_channel_label": "Dataset comercial precargado · sincronización DigiForms simulada",
                    "automatic_forms_results_api_enabled": True,
                    "automatic_forms_results_api_status": "demo_ready",
                    "automatic_forms_results_api_message": "La demo utiliza un dataset aislado con respuestas, georreferencias y formularios precargados. La sincronización se simula sin consumir servicios externos.",
                    "user_api_scope": "Entorno de demostración aislado: cosecha, malezas y plagas disponibles para recorrido comercial.",
                    "connection": {"company_id": company_id, "client_id": "DEMO-CLIENT", "api_user": "demo-api-user", "has_password": True, "connection_status": "demo_ready", "auto_sync_enabled": True},
                    "form_mappings": mappings,
                    "configured_form_types": [HARVEST_FORM_TYPE, PEST_WEED_FORM_TYPE],
                    "configured_forms": {HARVEST_FORM_TYPE: "DEMO-HARVEST-01", PEST_WEED_FORM_TYPE: "DEMO-PEST-01"},
                    "cursors": cursors,
                    "latest_import": latest,
                },
                "error": None,
            }
        connection = connection_for_company(db, company_id)
        api = _data_api_for_company(db, company_id)
        configured_types = _configured_form_types(db, company_id)
        api_enabled = api.is_configured and bool(configured_types)
        imports = [row for row in table(db, "sig_import_runs") if _same_company_or_legacy_user(row, company_id=company_id, user_id=user_id)]
        latest = max(imports, key=lambda row: str(row.get("created_at") or ""), default=None)
        cursors = [
            _cursor_public(row)
            for row in table(db, "sig_sync_cursors")
            if _same_company_or_legacy_user(row, company_id=company_id, user_id=user_id)
        ]
        mappings = safe_mappings(mappings_for_company(db, company_id))
    return {
        "data": {
            "company_id": company_id,
            "capture_platform": "DigiformsApp",
            "active_results_channel": DIGIFORMS_RESULTS_API_SOURCE if api_enabled else DIGIFORMS_EXCEL_SOURCE,
            "active_results_channel_label": "DigiForms Data API: carga inicial por fechas + incremental por ResponseId" if api_enabled else "Exportación Excel de DigiForms (respaldo)",
            "automatic_forms_results_api_enabled": api_enabled,
            "automatic_forms_results_api_status": "configured" if api_enabled else "missing_configuration",
            "automatic_forms_results_api_message": (
                "La API REST/JSON oficial está configurada para tu empresa. La primera carga usa fechas; las siguientes consultan respuestas incrementales por ResponseId. Cada respuesta conserva sus campos dinámicos y enlaces de imágenes."
                if api_enabled
                else "Configura la cuenta y al menos un FormId desde Extensiones → DigiForms. Mientras tanto, la carga Excel permanece disponible como respaldo."
            ),
            "user_api_scope": "User API administra usuarios. Data API sincroniza respuestas e imágenes para SIG Agrícola.",
            "connection": safe_connection(connection),
            "form_mappings": mappings,
            "configured_form_types": configured_types,
            "configured_forms": {
                HARVEST_FORM_TYPE: _form_id_for_type(db, company_id, HARVEST_FORM_TYPE) or None,
                PEST_WEED_FORM_TYPE: _form_id_for_type(db, company_id, PEST_WEED_FORM_TYPE) or None,
            },
            "cursors": cursors,
            "latest_import": latest,
        },
        "error": None,
    }


@router.post("/sync")
async def sync_from_digiforms_api(
    payload: Dict[str, Any] = Body(default_factory=dict),
    authorization: Optional[str] = Header(default=None),
):
    user = _require_user(authorization)
    user_id = str(user.get("id") or "")
    if is_commercial_demo_user(user):
        with LOCK:
            db = read_db()
            company_id = _company_id_for_user(db, user_id)
            rows = [row for row in table(db, "sig_import_runs") if _same_company_or_legacy_user(row, company_id=company_id, user_id=user_id)]
            rows.sort(key=lambda row: str(row.get("created_at") or ""), reverse=True)
        return {"data": {"runs": rows[:2], "errors": [], "demo": True}, "error": None, "message": "Dataset comercial DigiForms actualizado sin consumir servicios externos."}
    with LOCK:
        db = read_db()
        company_id = _company_id_for_user(db, user_id)
        if company_id and not extension_enabled_for(db, company_id, user_id):
            raise HTTPException(status_code=403, detail="DigiForms no está habilitado para tu empresa")
        form_types = _sync_types(db, company_id, payload)
    if not form_types:
        raise HTTPException(status_code=400, detail="Configura al menos un FormId desde Extensiones → DigiForms para sincronizar.")
    parcels = _sync_payload_parcels(payload, user_id, company_id)
    from_response_id = payload.get("from_response_id")
    if from_response_id not in (None, ""):
        try:
            from_response_id = int(from_response_id)
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail="from_response_id debe ser numérico.") from exc
    runs: List[Dict[str, Any]] = []
    errors: List[Dict[str, str]] = []
    for form_type in form_types:
        try:
            runs.append(await _sync_form_from_api(company_id=company_id, user_id=user_id, form_type=form_type, parcels=parcels, from_response_id=from_response_id))
        except HTTPException as exc:
            errors.append({"form_type": form_type, "message": str(exc.detail)})
    if errors and not runs:
        raise HTTPException(status_code=502, detail=" | ".join(item["message"] for item in errors))
    return {"data": {"runs": runs, "errors": errors}, "error": None, "message": "Sincronización DigiForms completada." if not errors else "Sincronización parcial completada."}


@router.get("/sync-cursors")
def list_sync_cursors(authorization: Optional[str] = Header(default=None)):
    user = _require_user(authorization)
    user_id = str(user.get("id") or "")
    with LOCK:
        db = read_db(); company_id = _company_id_for_user(db, user_id)
        rows = [_cursor_public(row) for row in table(db, "sig_sync_cursors") if _same_company_or_legacy_user(row, company_id=company_id, user_id=user_id)]
    return {"data": rows, "error": None, "count": len(rows)}


@router.post("/sync/test")
async def test_digiforms_data_api(payload: Dict[str, Any] = Body(default_factory=dict), authorization: Optional[str] = Header(default=None)):
    user = _require_user(authorization); user_id = str(user.get("id") or "")
    if is_commercial_demo_user(user):
        return {"data": {"ok": True, "records_received": 22, "form_id": "DEMO-HARVEST-01", "form_type": str(payload.get("form_type") or HARVEST_FORM_TYPE), "demo": True, "message": "Conexión simulada disponible para el recorrido comercial."}, "error": None}
    with LOCK:
        db = read_db(); company_id = _company_id_for_user(db, user_id)
        form_type = _safe_text(payload.get("form_type") or HARVEST_FORM_TYPE)
        form_id = _form_id_for_type(db, company_id, form_type)
        if not form_id:
            configured = _configured_form_types(db, company_id)
            if not configured: raise HTTPException(status_code=400, detail="No hay FormId configurado para probar DigiForms Data API.")
            form_type = configured[0]; form_id = _form_id_for_type(db, company_id, form_type)
        api = _data_api_for_company(db, company_id)
    if not api.is_configured: raise HTTPException(status_code=400, detail="DigiForms Data API no está configurada para esta empresa.")
    with LOCK:
        db = read_db()
        initial_start_date, initial_end_date = _initial_sync_dates_for_type(db, company_id, form_type)
    try:
        result = await api.test_results_connection(
            form_id,
            int(payload.get("from_response_id") or 0),
            start_date=_safe_text(payload.get("start_date")) or initial_start_date,
            end_date=_safe_text(payload.get("end_date")) or initial_end_date,
        )
    except DigiformsDataAPIError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    return {"data": {**result, "form_type": form_type}, "error": None}


def _active_company_actor(db: Dict[str, Any], company_id: str, connection: Dict[str, Any]) -> Optional[str]:
    active_user_ids = {str(user.get("id") or "") for user in db.get("users", []) if user.get("is_active", True) is not False and user.get("id")}
    configured_by = str(connection.get("configured_by_user_id") or "")
    if configured_by in active_user_ids: return configured_by
    for admin in table(db, "admin_users"):
        candidate = str(admin.get("user_id") or "")
        if str(admin.get("company_id") or "") == company_id and admin.get("is_active", True) is not False and candidate in active_user_ids:
            return candidate
    return None


@router.post("/sync/cron")
async def cron_sync_from_digiforms_api(x_dataris_cron_secret: Optional[str] = Header(default=None, alias="X-Dataris-Cron-Secret")):
    expected = _safe_text(settings.DIGIFORMS_SYNC_CRON_SECRET)
    if not expected or x_dataris_cron_secret != expected: raise HTTPException(status_code=401, detail="Cron secret inválido.")
    with LOCK:
        db = read_db()
        targets: List[Tuple[str, str, List[str], List[Dict[str, Any]]]] = []
        for connection in table(db, "digiforms_connections"):
            company_id = str(connection.get("company_id") or "")
            if not company_id or connection.get("auto_sync_enabled", True) is False: continue
            if not extension_enabled_for(db, company_id, None): continue
            actor = _active_company_actor(db, company_id, connection)
            if not actor: continue
            form_types = _configured_form_types(db, company_id)
            if not form_types: continue
            targets.append((company_id, actor, form_types, _stored_parcels(db, actor, company_id)))
    results: List[Dict[str, Any]] = []; errors: List[Dict[str, str]] = []
    for company_id, user_id, form_types, parcels in targets:
        for form_type in form_types:
            try:
                run = await _sync_form_from_api(company_id=company_id, user_id=user_id, form_type=form_type, parcels=parcels)
                results.append({"company_id": company_id, "user_id": user_id, "form_type": form_type, "run": run})
            except HTTPException as exc:
                errors.append({"company_id": company_id, "user_id": user_id, "form_type": form_type, "message": str(exc.detail)})
    return {"data": {"runs": results, "errors": errors, "companies_processed": len(targets)}, "error": None}


@router.get("/raw-submissions")
def list_raw_submissions(form_type: Optional[str] = None, authorization: Optional[str] = Header(default=None)):
    """Expose preserved dynamic DigiForms payloads for diagnostics and future mappings."""
    user = _require_user(authorization); user_id = str(user.get("id") or "")
    with LOCK:
        db = read_db(); company_id = _company_id_for_user(db, user_id)
        rows = [dict(row) for row in table(db, "sig_digiforms_raw_submissions") if _same_company_or_legacy_user(row, company_id=company_id, user_id=user_id)]
        if form_type: rows = [row for row in rows if row.get("form_type") == form_type]
        rows = sorted(rows, key=lambda row: str(row.get("updated_at") or row.get("created_at") or ""), reverse=True)[:500]
    return {"data": rows, "error": None, "count": len(rows)}


@router.get("/imports")
def list_imports(form_type: Optional[str] = None, authorization: Optional[str] = Header(default=None)):
    user = _require_user(authorization); user_id = str(user.get("id") or "")
    with LOCK:
        db = read_db(); company_id = _company_id_for_user(db, user_id)
        rows = [row for row in table(db, "sig_import_runs") if _same_company_or_legacy_user(row, company_id=company_id, user_id=user_id)]
        if form_type: rows = [row for row in rows if row.get("form_type") == form_type]
        rows = sorted(rows, key=lambda row: str(row.get("created_at") or ""), reverse=True)[:100]
    return {"data": rows, "error": None, "count": len(rows)}


MAX_PLANNING_SHAPEFILE_BYTES = 30 * 1024 * 1024


@router.post("/planning/preview")
async def preview_planning_upload(
    shapefile: UploadFile = File(...),
    excel: UploadFile = File(...),
    id_field: Optional[str] = Form(default=None),
    authorization: Optional[str] = Header(default=None),
):
    """Cruza un shapefile de lotes (.zip) con un Excel de planificación por ID.

    No persiste nada: devuelve el GeoJSON combinado y los campos categóricos
    disponibles para que el frontend arme la simbología (color por atributo)
    y la leyenda, similar a un "categorized renderer" de QGIS/ArcGIS.
    """
    _require_user(authorization)

    shapefile_bytes = await shapefile.read()
    if not shapefile_bytes:
        raise HTTPException(status_code=400, detail="El archivo de lotes (shapefile) está vacío.")
    if len(shapefile_bytes) > MAX_PLANNING_SHAPEFILE_BYTES:
        raise HTTPException(status_code=413, detail="El shapefile supera el límite de 30 MB.")

    excel_bytes = await _read_upload(excel)

    try:
        result = await run_in_threadpool(
            process_planning_upload, shapefile_bytes, excel_bytes, id_field or None
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error procesando los archivos: {exc}")

    return {"data": result, "error": None}


@router.post("/imports/harvest")
async def import_harvest(file: UploadFile = File(...), parcels_json: str = Form("[]"), authorization: Optional[str] = Header(default=None)):
    user = _require_user(authorization); user_id = str(user.get("id") or ""); content = await _read_upload(file); parcels = _parse_parcels(parcels_json)
    with LOCK:
        db = read_db(); company_id = _company_id_for_user(db, user_id)
        run = _import_records(db=db, user_id=user_id, company_id=company_id, form_type=HARVEST_FORM_TYPE, filename=file.filename or "digiforms-cosecha.xlsx", content=content, parcels=parcels); write_db(db)
    return {"data": run, "error": None, "message": "Exportación de cosecha importada y persistida correctamente."}


@router.post("/imports/pest-weed")
async def import_pest_weed(file: UploadFile = File(...), parcels_json: str = Form("[]"), authorization: Optional[str] = Header(default=None)):
    user = _require_user(authorization); user_id = str(user.get("id") or ""); content = await _read_upload(file); parcels = _parse_parcels(parcels_json)
    with LOCK:
        db = read_db(); company_id = _company_id_for_user(db, user_id)
        run = _import_records(db=db, user_id=user_id, company_id=company_id, form_type=PEST_WEED_FORM_TYPE, filename=file.filename or "digiforms-malezas-plagas.xlsx", content=content, parcels=parcels); write_db(db)
    return {"data": run, "error": None, "message": "Exportación de malezas y plagas importada y persistida correctamente."}


@router.get("/harvest-records")
def list_harvest_records(include_non_approved: bool = Query(default=False), authorization: Optional[str] = Header(default=None)):
    user = _require_user(authorization); user_id = str(user.get("id") or "")
    with LOCK:
        db = read_db(); company_id = _company_id_for_user(db, user_id)
        rows = [_safe_record(row) for row in table(db, "sig_harvest_records") if _same_company_or_legacy_user(row, company_id=company_id, user_id=user_id) and (include_non_approved or _safe_text(row.get("digiforms_state") or "0") == "0")]
        rows.sort(key=lambda row: str(row.get("updated_at") or row.get("created_at") or ""), reverse=True)
    return {"data": rows, "error": None, "count": len(rows)}


@router.get("/pest-weed-records")
def list_pest_weed_records(include_non_approved: bool = Query(default=False), authorization: Optional[str] = Header(default=None)):
    user = _require_user(authorization); user_id = str(user.get("id") or "")
    with LOCK:
        db = read_db(); company_id = _company_id_for_user(db, user_id)
        rows = [_safe_record(row) for row in table(db, "sig_pest_weed_records") if _same_company_or_legacy_user(row, company_id=company_id, user_id=user_id) and (include_non_approved or _safe_text(row.get("digiforms_state") or "0") == "0")]
        rows.sort(key=lambda row: str(row.get("updated_at") or row.get("created_at") or ""), reverse=True)
    return {"data": rows, "error": None, "count": len(rows)}


@router.get("/harvest-overrides")
def list_harvest_overrides(authorization: Optional[str] = Header(default=None)):
    user = _require_user(authorization); user_id = str(user.get("id") or "")
    with LOCK:
        db = read_db(); company_id = _company_id_for_user(db, user_id)
        rows = [dict(row) for row in table(db, "sig_harvest_overrides") if _same_company_or_legacy_user(row, company_id=company_id, user_id=user_id)]
        rows.sort(key=lambda row: str(row.get("updated_at") or ""), reverse=True)
    return {"data": rows, "error": None, "count": len(rows)}


@router.put("/harvest-overrides")
def upsert_harvest_override(payload: Dict[str, Any] = Body(default_factory=dict), authorization: Optional[str] = Header(default=None)):
    user = _require_user(authorization); user_id = str(user.get("id") or ""); feature_key = _safe_text(payload.get("feature_key")); parcel_id = _safe_text(payload.get("parcel_id"))
    if not feature_key or not parcel_id: raise HTTPException(status_code=400, detail="feature_key y parcel_id son obligatorios.")
    timestamp = now()
    with LOCK:
        db = read_db(); company_id = _company_id_for_user(db, user_id); rows = table(db, "sig_harvest_overrides")
        existing = next((row for row in rows if _same_company_or_legacy_user(row, company_id=company_id, user_id=user_id) and row.get("feature_key") == feature_key), None)
        values = {"user_id": user_id, "company_id": company_id, "feature_key": feature_key, "parcel_id": parcel_id, "estado": _canonical_harvest_state(payload.get("estado")), "metodo": _safe_text(payload.get("metodo")), "fecha": _safe_text(payload.get("fecha")), "terminal": _safe_text(payload.get("terminal")), "zona": _safe_text(payload.get("zona")), "updated_at": timestamp}
        if existing: existing.update(values); row = existing
        else: row = {"id": str(uuid.uuid4()), "created_at": timestamp, **values}; rows.append(row)
        write_db(db)
    return {"data": row, "error": None, "message": "Corrección manual guardada."}


@router.delete("/harvest-overrides/{feature_key:path}")
def delete_harvest_override(feature_key: str, authorization: Optional[str] = Header(default=None)):
    user = _require_user(authorization); user_id = str(user.get("id") or "")
    with LOCK:
        db = read_db(); company_id = _company_id_for_user(db, user_id); rows = table(db, "sig_harvest_overrides"); before = len(rows)
        db["tables"]["sig_harvest_overrides"] = [row for row in rows if not (_same_company_or_legacy_user(row, company_id=company_id, user_id=user_id) and row.get("feature_key") == feature_key)]
        deleted = before != len(db["tables"]["sig_harvest_overrides"]); write_db(db)
    if not deleted: raise HTTPException(status_code=404, detail="No se encontró la corrección manual.")
    return {"data": {"ok": True}, "error": None}
