"""Exportación de la bitácora al formato de hoja de cálculo de siempre.

Este archivo es, en la práctica, la palanca de adopción del módulo: el reporte
semanal que hoy se manda por correo se sigue pudiendo mandar igual, pero sale
de un botón en vez de una transcripción a mano. Nadie tiene que renunciar a su
formato para empezar a capturar en campo.

Se conserva el orden y los rótulos de la hoja original (diez bloques de costos,
resultados técnicos, resumen y sensibilidad) para que quien la reciba no note
la diferencia.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.modules.field_log.models import CATEGORY_LABELS, LOG_CATEGORIES

_TITLE_FILL = PatternFill("solid", fgColor="1F3B2C")
_SECTION_FILL = PatternFill("solid", fgColor="DCE9DF")
_HEADER_FILL = PatternFill("solid", fgColor="EFEFEF")
_TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
_SECTION_FONT = Font(bold=True, size=11)
_BOLD = Font(bold=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_MONEY = '#,##0.00'
_DECIMAL = '#,##0.00'


def _write_row(ws, row: int, values: list[Any], *, bold: bool = False, fill=None, number_format: str | None = None):
    for offset, value in enumerate(values):
        cell = ws.cell(row=row, column=2 + offset, value=value)
        if bold:
            cell.font = _BOLD
        if fill is not None:
            cell.fill = fill
        if number_format and isinstance(value, (int, float)):
            cell.number_format = number_format
    return row + 1


def _section(ws, row: int, title: str) -> int:
    cell = ws.cell(row=row, column=2, value=title)
    cell.font = _SECTION_FONT
    for column in range(2, 11):
        ws.cell(row=row, column=column).fill = _SECTION_FILL
    return row + 1


def _fmt(value: Any) -> Any:
    """None se escribe como el '#DIV/0!' que ya conocen de la hoja original."""
    if value is None:
        return "—"
    return value


def build_workbook(summary: dict[str, Any], entries: list[dict[str, Any]], sensitivity: dict[str, Any]) -> bytes:
    cycle = summary["cycle"]
    kpis = summary["kpis"]
    economics = kpis["economics"]
    sustainability = kpis["sustainability"]
    costs = kpis["costs"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Registro"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    for letter in ("C", "D", "E", "F", "G", "H", "I", "J"):
        ws.column_dimensions[letter].width = 17

    row = 2
    title = ws.cell(row=row, column=2, value=f"BITÁCORA CICLO: {cycle.get('name', '')}")
    title.font = _TITLE_FONT
    for column in range(2, 11):
        ws.cell(row=row, column=column).fill = _TITLE_FILL
    row += 2

    parcel = cycle.get("parcel_name") or ""
    row = _write_row(ws, row, ["Parcela", parcel, "Cultivo", cycle.get("crop_type") or ""])
    row = _write_row(
        ws,
        row,
        [
            "Variedad",
            cycle.get("variety") or "",
            "Superficie (ha)",
            cycle.get("area_ha"),
        ],
    )
    row = _write_row(
        ws,
        row,
        [
            "Siembra",
            str(cycle.get("planting_date") or ""),
            "Cosecha",
            str(cycle.get("harvest_date") or ""),
        ],
    )
    row += 1

    # ------------------------------------------------------ bloque económico
    row = _section(ws, row, "RESULTADOS ECONÓMICOS")
    row = _write_row(
        ws,
        row,
        ["RENDIMIENTO POR HA", _fmt(economics["yield_ton_ha"]), "PRECIO DE VENTA", _fmt(economics["price_per_ton"])],
        number_format=_DECIMAL,
    )
    row = _write_row(
        ws,
        row,
        ["INGRESO", _fmt(economics["revenue_per_ha"]), "INVERSIÓN POR HA", _fmt(economics["investment_per_ha"])],
        number_format=_MONEY,
    )
    row = _write_row(
        ws,
        row,
        ["UTILIDAD POR HA", _fmt(economics["profit_per_ha"]), "REL B/C", _fmt(economics["benefit_cost_ratio"])],
        number_format=_MONEY,
    )
    row = _write_row(
        ws,
        row,
        [
            "PUNTO DE EQUILIBRIO (TON/HA)",
            _fmt(economics["break_even_ton_ha"]),
            "COSTO UNITARIO ($/TON)",
            _fmt(economics["unit_cost_per_ton"]),
        ],
        number_format=_DECIMAL,
    )
    row += 1

    # ------------------------------------------------------ resultados técnicos
    row = _section(ws, row, "RESULTADOS TÉCNICOS")
    technical = [
        ("HUELLA HÍDRICA", "CONSUMO DE AGUA", sustainability["water_footprint_m3_per_ton"], "M3/TON"),
        ("ENERGÍA", "ELECTRICIDAD", sustainability["energy_kwh_per_ton"], "KWH/TON"),
        ("", "DIESEL", sustainability["diesel_l_per_ha"], "L/HA"),
        ("INSUMOS", "AGROQUÍMICOS", sustainability["insecticide_ia_g_per_ton"], "GRAMOS DE I.A. DE INSECTICIDA/TON"),
        ("", "", sustainability["herbicide_ia_g_per_ton"], "GRAMOS DE I.A. DE HERBICIDA/TON"),
        ("", "", sustainability["fungicide_ia_g_per_ton"], "GRAMOS DE I.A. DE FUNGICIDA/TON"),
        ("", "FERTILIZANTE", sustainability["nitrogen_kg_per_ton"], "KG DE N/TON"),
    ]
    for group, concept, value, unit in technical:
        row = _write_row(ws, row, [group, concept, _fmt(value), unit], number_format=_DECIMAL)
    row += 1

    # ------------------------------------------------------ bloques de costos
    row = _section(ws, row, "DESCRIPCIÓN DE CONCEPTOS")
    header = ["DESCRIPCIÓN", "UNIDADES", "CANTIDAD", "C.U.", "COSTO / HA ($)", "FECHA DE REALIZACIÓN", "OBSERVACIONES", "DATOS ADICIONALES"]
    row = _write_row(ws, row, header, bold=True, fill=_HEADER_FILL)

    by_category: dict[str, list[dict[str, Any]]] = {key: [] for key in LOG_CATEGORIES}
    for entry in entries:
        by_category.setdefault(entry.get("category"), []).append(entry)

    category_costs = {item["category"]: item for item in costs["categories"]}

    # El formato en papel cierra cada bloque con sus propias líneas: los litros
    # de diesel del acondicionamiento, los kWh y m³ de los riegos, la fórmula
    # N-P-K, los gramos de i.a. Se reproducen aquí para que el archivo exportado
    # se lea igual que el original y no como un simple listado de labores.
    template = summary.get("template") or {}
    blocks = kpis.get("blocks") or {}
    cycle_attributes = (cycle.get("attributes") or {}) if isinstance(cycle, dict) else {}
    declarations = {
        item.get("key"): item for item in (template.get("categories") or []) if item.get("key")
    }

    for category in LOG_CATEGORIES:
        block = by_category.get(category) or []
        subtotal = category_costs.get(category, {}).get("cost_per_ha", 0.0)
        row = _write_row(
            ws,
            row,
            [CATEGORY_LABELS[category], "", "", "", subtotal],
            bold=True,
            fill=_SECTION_FILL,
            number_format=_MONEY,
        )
        for entry in sorted(block, key=lambda item: (item.get("performed_at") is None, item.get("performed_at"))):
            extra = entry.get("data") or {}
            extra_text = ", ".join(f"{key}: {value}" for key, value in extra.items() if value not in (None, ""))
            row = _write_row(
                ws,
                row,
                [
                    entry.get("description"),
                    entry.get("unit"),
                    entry.get("quantity"),
                    entry.get("unit_cost"),
                    entry.get("cost_per_ha"),
                    str(entry.get("performed_at") or ""),
                    entry.get("observations") or "",
                    extra_text,
                ],
                number_format=_DECIMAL,
            )

        declaration = declarations.get(category) or {}
        block_totals = blocks.get(category) or {}
        for total in declaration.get("totals") or []:
            value = block_totals.get(total.get("key"))
            if value in (None, ""):
                continue
            row = _write_row(ws, row, [total.get("label"), "", value], number_format=_DECIMAL)

        for attribute in declaration.get("attributes") or []:
            value = cycle_attributes.get(attribute.get("name"))
            if value in (None, ""):
                continue
            label = attribute.get("label")
            unit = attribute.get("unit")
            row = _write_row(ws, row, [f"{label} ({unit})" if unit else label, "", value])

    row = _write_row(
        ws,
        row,
        ["SUMA DE CONCEPTOS", "", "", "", costs["total_cost_per_ha"]],
        bold=True,
        number_format=_MONEY,
    )
    row += 1

    # ------------------------------------------------------ resumen de costos
    row = _section(ws, row, "RESUMEN DE COSTOS")
    row = _write_row(ws, row, ["CONCEPTO", "COSTO", "%"], bold=True, fill=_HEADER_FILL)
    for item in costs["categories"]:
        row = _write_row(
            ws,
            row,
            [item["label"], item["cost_per_ha"], _fmt(item["percentage"])],
            number_format=_DECIMAL,
        )
    row = _write_row(ws, row, ["TOTAL", costs["total_cost_per_ha"], 100 if costs["total_cost_per_ha"] else "—"], bold=True, number_format=_DECIMAL)
    row += 1

    # ------------------------------------------------------ fenología
    phenology = summary.get("phenology") or []
    if phenology:
        row = _section(ws, row, "REGISTRO DE FENOLOGÍA")
        row = _write_row(ws, row, ["Etapa", "Fecha", "Observaciones"], bold=True, fill=_HEADER_FILL)
        for record in phenology:
            label = getattr(record, "stage_label", None) or getattr(record, "stage_code", "")
            observed = getattr(record, "observed_at", None)
            notes = getattr(record, "observations", None)
            row = _write_row(ws, row, [label, str(observed or ""), notes or ""])
        row += 1

    # ------------------------------------------------------ sensibilidad
    row = _section(ws, row, "ANÁLISIS DE SENSIBILIDAD")
    row = _write_row(ws, row, ["", "PRECIO DE VENTA"], bold=True)
    row = _write_row(ws, row, ["REND. (TON/HA)", *sensitivity["price_axis"]], bold=True, fill=_HEADER_FILL)
    for matrix_row in sensitivity["rows"]:
        row = _write_row(
            ws,
            row,
            [matrix_row["yield_ton_ha"], *[cell["profit_per_ha"] for cell in matrix_row["cells"]]],
            number_format=_MONEY,
        )

    # ------------------------------------------------------ verificación
    verification = summary.get("verification") or {}
    if verification.get("total"):
        row += 1
        row = _section(ws, row, "TRAZABILIDAD DE CAPTURA")
        row = _write_row(ws, row, ["Registros verificados en campo (GPS)", verification.get("verified", 0)])
        row = _write_row(ws, row, ["Registros fuera del polígono", verification.get("outside", 0)])
        row = _write_row(ws, row, ["Registros sin ubicación", verification.get("unknown", 0)])

    for column in range(2, 11):
        ws.cell(row=1, column=column).alignment = Alignment(horizontal="center")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def default_file_name(cycle_name: str) -> str:
    safe = "".join(ch if ch.isalnum() or ch in "-_ " else "_" for ch in (cycle_name or "bitacora")).strip()
    return f"Bitacora - {safe or 'ciclo'}.xlsx"
