"""Importación del histórico desde la hoja de cálculo.

Permite arrancar un ciclo con lo que ya está capturado en Excel en lugar de
volver a teclearlo, que es la condición para que alguien acepte cambiar de
herramienta a mitad de temporada.

El lector es deliberadamente tolerante: estas hojas se editan a mano durante
meses, así que se localizan los bloques por su rótulo ("1. ACONDICIONAMIENTO")
en lugar de asumir números de fila fijos, y cualquier fila que no se entienda
se ignora en vez de abortar la importación completa.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.modules.field_log.models import LOG_CATEGORIES

# El número que precede al rótulo del bloque es el índice de la categoría en el
# orden canónico de la bitácora.
_BLOCK_PATTERN = re.compile(r"^\s*(\d{1,2})\s*\.\s*(.+)$")

_DESCRIPTION_COL = 2
_UNIT_COL = 4
_QUANTITY_COL = 5
_UNIT_COST_COL = 6
_COST_COL = 7
_EXTRA_COLS = range(8, 13)

_PHENOLOGY_STAGE_COL = 13
_PHENOLOGY_DATE_COL = 14
_PHENOLOGY_NOTES_COL = 15

_SKIP_ROW_TOKENS = {
    "descripcion de conceptos",
    "descripción de conceptos",
    "total",
    "concepto",
    "formula",
    "fórmula",
    "descripción",
    "descripcion",
}

# Rótulos que cierran la zona de labores. Todo lo que viene después son
# resúmenes y matrices —incluido el bloque de sensibilidad, cuyas filas son
# números que se leerían como labores fantasma— y ya no se importa. Es lo que
# hace que reimportar un archivo exportado por Dataris devuelva exactamente las
# mismas labores y no unas cuantas de más.
_END_OF_ENTRIES_TOKENS = {
    "suma de conceptos",
    "resumen de costos",
    "analisis de sensibilidad",
    "análisis de sensibilidad",
    "trazabilidad de captura",
    "registro de fenologia",
    "registro de fenología",
}


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = _text(value).replace(",", "").replace("$", "")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _category_from_label(label: str) -> str | None:
    match = _BLOCK_PATTERN.match(label)
    if not match:
        return None
    index = int(match.group(1))
    if 1 <= index <= len(LOG_CATEGORIES):
        return LOG_CATEGORIES[index - 1]
    return None


def parse_workbook(content: bytes) -> dict[str, Any]:
    """Extrae ciclo, labores y fenología de una bitácora en Excel."""
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.worksheets[0]

    entries: list[dict[str, Any]] = []
    phenology: list[dict[str, Any]] = []
    cycle: dict[str, Any] = {}
    warnings: list[str] = []

    current_category: str | None = None
    entries_closed = False
    max_row = min(sheet.max_row, 2000)

    for row_index in range(1, max_row + 1):
        first = _text(sheet.cell(row=row_index, column=_DESCRIPTION_COL).value)

        if first.lower() in _END_OF_ENTRIES_TOKENS:
            entries_closed = True
            current_category = None

        if first.upper().startswith("BITÁCORA CICLO") or first.upper().startswith("BITACORA CICLO"):
            cycle["name"] = first.split(":", 1)[-1].strip() or first

        # Cabecera económica: los rótulos y sus valores están en la fila siguiente.
        if first.upper().startswith("RENDIMIENTO POR HA"):
            values_row = row_index + 1
            cycle.setdefault("actual_yield_ton_ha", _number(sheet.cell(row=values_row, column=2).value))
            cycle.setdefault("target_price_per_ton", _number(sheet.cell(row=values_row, column=4).value))

        stage = _text(sheet.cell(row=row_index, column=_PHENOLOGY_STAGE_COL).value)
        if stage and stage.lower() not in {"etapa", "labor"}:
            observed = _as_date(sheet.cell(row=row_index, column=_PHENOLOGY_DATE_COL).value)
            notes = _text(sheet.cell(row=row_index, column=_PHENOLOGY_NOTES_COL).value)
            if observed or notes:
                phenology.append(
                    {
                        "stage_code": stage.lower().replace(" ", "_")[:40],
                        "stage_label": stage,
                        "observed_at": observed,
                        "observations": notes or None,
                    }
                )

        detected = _category_from_label(first)
        if detected:
            # Tras el cierre, un "1. Acondicionamiento" es la fila del resumen
            # de costos, no la cabecera de un bloque de labores.
            current_category = None if entries_closed else detected
            continue

        if not current_category or not first:
            continue
        if first.lower() in _SKIP_ROW_TOKENS:
            continue

        quantity = _number(sheet.cell(row=row_index, column=_QUANTITY_COL).value)
        unit_cost = _number(sheet.cell(row=row_index, column=_UNIT_COST_COL).value)
        cost = _number(sheet.cell(row=row_index, column=_COST_COL).value)

        # Una fila sin ningún número es un rótulo suelto del bloque, no una labor.
        if quantity is None and unit_cost is None and not cost:
            continue

        performed_at: date | None = None
        observations: str | None = None
        extra: dict[str, Any] = {}

        for column in _EXTRA_COLS:
            value = sheet.cell(row=row_index, column=column).value
            parsed_date = _as_date(value)
            if parsed_date and performed_at is None:
                performed_at = parsed_date
                continue
            number = _number(value)
            if number is not None:
                extra[f"col_{column}"] = number
                continue
            text = _text(value)
            if text and observations is None:
                observations = text

        entries.append(
            {
                "category": current_category,
                "description": first[:400],
                "unit": _text(sheet.cell(row=row_index, column=_UNIT_COL).value) or None,
                "quantity": quantity,
                "unit_cost": unit_cost,
                "cost_per_ha": cost if cost is not None else None,
                "performed_at": performed_at,
                "observations": observations,
                "data": extra or None,
                "source": "import",
            }
        )

    if not entries:
        warnings.append(
            "No se reconoció ninguna labor. Verifica que la hoja tenga los bloques "
            "numerados (1. ACONDICIONAMIENTO, 2. SIEMBRA, …) en la segunda columna."
        )

    return {
        "cycle": cycle,
        "entries": entries,
        "phenology": phenology,
        "warnings": warnings,
    }
